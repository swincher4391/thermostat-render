import subprocess
import sys

# Install required libraries
try:
    import requests
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "openpyxl"])
    import requests
    import openpyxl

import os

# URLs for KY PSC Case 2021-00214 WNA documents
urls = {
    "Residential_WNA_Customer": "https://psc.ky.gov/pscecf/2021-00214/regulatory.support%40atmosenergy.com/09162021033037/Staff_3-09_Att1_-_Residential_WNA_Customer.xlsx",
    "Bill_Cycle_Normals": "https://psc.ky.gov/pscecf/2021-00214/regulatory.support%40atmosenergy.com/09162021033037/Staff_3-09_Att3_-_10_Yr_and_20_Yr_Bill_Cycle_Normals_Ending_Mar21.xlsx"
}

output_dir = r"C:\dev\Budget\Atmos"

print("=" * 70)
print("Downloading KY PSC WNA Data Files")
print("=" * 70)

for name, url in urls.items():
    filename = os.path.join(output_dir, f"{name}.xlsx")
    print(f"\nDownloading: {name}")
    print(f"  URL: {url[:80]}...")

    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()

        with open(filename, "wb") as f:
            f.write(response.content)
        print(f"  Saved to: {filename}")
        print(f"  Size: {len(response.content):,} bytes")
    except Exception as e:
        print(f"  ERROR: {e}")

print("\n" + "=" * 70)
print("Extracting WNA Parameters")
print("=" * 70)

# Read Residential WNA Customer file
wna_file = os.path.join(output_dir, "Residential_WNA_Customer.xlsx")
if os.path.exists(wna_file):
    print(f"\nReading: {wna_file}")
    wb = openpyxl.load_workbook(wna_file)

    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]

        # Print first 30 rows to understand structure
        for row_num, row in enumerate(ws.iter_rows(max_row=30, values_only=True), 1):
            # Filter out completely empty rows
            if any(cell is not None for cell in row):
                row_str = " | ".join(str(cell)[:30] if cell else "" for cell in row[:10])
                print(f"  Row {row_num}: {row_str}")

# Read Bill Cycle Normals file
normals_file = os.path.join(output_dir, "Bill_Cycle_Normals.xlsx")
if os.path.exists(normals_file):
    print(f"\n\nReading: {normals_file}")
    wb = openpyxl.load_workbook(normals_file)

    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]

        # Print first 30 rows
        for row_num, row in enumerate(ws.iter_rows(max_row=30, values_only=True), 1):
            if any(cell is not None for cell in row):
                row_str = " | ".join(str(cell)[:25] if cell else "" for cell in row[:12])
                print(f"  Row {row_num}: {row_str}")

print("\n" + "=" * 70)
print("Done!")
